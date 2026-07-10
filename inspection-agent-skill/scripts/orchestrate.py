#!/usr/bin/env python
"""
摄像头巡检 — 独立编排脚本

平台凭证已封装在技能包内部 config/platform.json，开箱即用。
用户只需提供：摄像头清单 Excel 路径、截图保存目录。
品牌文本（飞书卡片标题/页脚、启动横幅等）从 config/branding.json 加载，
切换业务场景时只需修改 branding.json，无需改动此脚本。

用法：
  python scripts/orchestrate.py                                          # 使用 config/platform.json
  python scripts/orchestrate.py --excel-path ./清单.xlsx --save-dir ./screenshots
  python scripts/orchestrate.py --config my.json                         # 完全自定义配置
  python scripts/orchestrate.py --test-single CAM-001                    # 单路测试
"""
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime

# 从同目录导入截图模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from camera_capture import capture_single


# ─── 品牌配置加载 ──────────────────────────────────────────

def load_branding_config(merged_config: dict = None) -> dict:
    """
    加载品牌配置，优先级：
      1) 合并后的主配置里若自带 "branding" 子对象（港区等场景通过 --config 传入）→ 优先使用
      2) 否则读取技能包 config/branding.json（环保默认）
    切换业务场景只需改对应 branding，无需改代码。
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    branding_path = os.path.join(script_dir, "..", "config", "branding.json")

    defaults = {
        "project_label": "",
        "feishu_card_title": "摄像头巡检报告",
        "feishu_card_footer": "巡检助手 · 自动生成",
        "inspection_start_banner": "摄像头巡检",
        "region_label": "",          # 快照Excel的『所属区域』列 / 文件标题前缀
    }

    # 1) 主配置自带 branding（港区场景）
    if isinstance(merged_config, dict) and isinstance(merged_config.get("branding"), dict):
        b = defaults.copy()
        b.update(merged_config["branding"])
        return b

    # 2) 技能包 branding.json（环保默认）
    if os.path.exists(branding_path):
        try:
            with open(branding_path, "r", encoding="utf-8") as f:
                branding = json.load(f)
            for k, v in defaults.items():
                branding.setdefault(k, v)
            return branding
        except Exception as e:
            print(f"[品牌配置] 读取 branding.json 失败: {e}，使用默认值")

    return defaults


# ─── 配置加载 ──────────────────────────────────────────

def load_platform_config() -> dict:
    """
    加载技能包本地 config/platform.json（兜底配置）。
    CLI 参数（--host/--app-key/--app-secret 等）优先级更高，可完全覆盖此配置。
    如果用户通过 CLI 提供了所有必填参数，则无需依赖此文件。
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    local_config = os.path.join(script_dir, "..", "config", "platform.json")

    config = {}
    if os.path.exists(local_config):
        with open(local_config, "r", encoding="utf-8") as f:
            config = json.load(f)
        print(f"[配置] 从技能包本地加载: {local_config}")
    else:
        print("[配置] 技能包本地配置不存在，使用环境变量")
        config["host"] = os.environ.get("HIK_HOST", "")
        config["app_key"] = os.environ.get("HIK_APP_KEY", "")
        config["app_secret"] = os.environ.get("HIK_APP_SECRET", "")
        config["feishu_webhook_url"] = os.environ.get("FEISHU_WEBHOOK_URL", "")

    config.setdefault("port", 443)
    config.setdefault("api_path", "/artemis")
    config.setdefault("timeout", 10)
    config.setdefault("retry_count", 2)
    config.setdefault("max_workers", 30)
    config.setdefault("save_dir", "./screenshots")
    config.setdefault("feishu_webhook_url", "")

    return config


# ─── Step 1: 读取摄像头清单 ─────────────────────────────
# 编码约定：所有文本文件读写统一使用 UTF-8 编码。
# Hermes 调度引擎在某些环境下可能默认使用 GBK 编码读取文件，
# 因此此处显式指定 encoding='utf-8' 以覆盖默认行为。

def parse_camera_excel(excel_path: str) -> list:
    """
    解析摄像头清单，支持 .xlsx / .csv 两种格式，统一使用 UTF-8 编码。

    Excel 列定义：
      A: 摄像头编码     B: 摄像头名称     C: 是否更新到台账  (是=写入Word台账)
      D 起: 每日巡检状态 (系统自动填入，按日追加)

    - .xlsx：通过 openpyxl 读取（二进制格式，编码由库内部处理）
    - .csv： 通过 csv 模块读取，显式指定 encoding='utf-8-sig'（兼容 BOM 头）
    """
    def _is_enabled(val) -> bool:
        if val is None:
            return False
        s = str(val).strip().lower()
        return s in ("是", "true", "1", "y", "yes", "√", "✓")

    ext = os.path.splitext(excel_path)[1].lower()

    # ── CSV 格式（UTF-8 编码） ──
    if ext == ".csv":
        import csv
        cameras = []
        with open(excel_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None)  # 跳过表头
            for row in reader:
                if len(row) >= 2 and row[0] and row[1]:
                    cameras.append({
                        "name": str(row[1]).strip(),
                        "code": str(row[0]).strip(),
                        "ledger_enabled": _is_enabled(row[2]) if len(row) > 2 else False,
                    })
        return cameras

    # ── XLSX 格式（默认） ──
    try:
        import openpyxl
    except ImportError:
        print("ERROR: 请安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    cameras = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, name = row[0], row[1]
        if name and code:
            cameras.append({
                "name": str(name).strip(),
                "code": str(code).strip(),
                "ledger_enabled": _is_enabled(row[2]) if len(row) > 2 else False,
            })
    return cameras


# ─── Step 2: 并发截图 ──────────────────────────────────

def run_single_camera(cam: dict, config: dict) -> dict:
    """对单个摄像头执行截图"""
    result = capture_single(
        host=config["host"],
        port=int(config.get("port", 443)),
        app_key=config["app_key"],
        app_secret=config["app_secret"],
        camera_index_code=cam["code"],
        camera_name=cam["name"],
        save_dir=config.get("save_dir", "./screenshots"),
        timeout=int(config.get("timeout", 10)),
        retry_count=int(config.get("retry_count", 2)),
        api_path=config.get("api_path", "/artemis"),
    )
    result["cameraName"] = cam["name"]
    result["cameraCode"] = cam["code"]
    result["ledgerEnabled"] = cam.get("ledger_enabled", False)
    return result


def execute_concurrent_inspection(cameras: list, config: dict) -> list:
    """并发执行所有摄像头截图，同步屏障等待全部完成"""
    all_results = {}
    camera_count = len(cameras)
    max_workers = min(int(config.get("max_workers", 30)), camera_count)
    print(f"  并发数: {max_workers} (共 {camera_count} 路)")

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_map = {pool.submit(run_single_camera, c, config): c for c in cameras}

        completed = 0
        for future in as_completed(future_map):
            cam = future_map[future]
            code = cam["code"]
            try:
                result = future.result(timeout=45)
                all_results[code] = result
            except Exception as e:
                all_results[code] = {
                    "status": "offline", "cameraName": cam["name"], "cameraCode": code,
                    "screenshotPath": None, "qualityScore": 0.0, "errorMsg": f"线程异常: {e}",
                }
            completed += 1

    # 按原始顺序输出
    ordered = []
    for cam in cameras:
        r = all_results.get(cam["code"], {
            "status": "offline", "cameraName": cam["name"], "cameraCode": cam["code"],
            "screenshotPath": None, "qualityScore": 0.0, "errorMsg": "未收到结果",
        })
        ordered.append(r)

    print(f"  完成 {completed}/{camera_count} 路")
    return ordered


# ─── Step 3: 汇总结果 ─────────────────────────────────

def summarize(results: list) -> dict:
    online = [r for r in results if r["status"] == "online"]
    offline = [r for r in results if r["status"] == "offline"]
    abnormal = [r for r in results if r["status"] == "abnormal"]
    total = len(results)

    return {
        "total": total,
        "online": len(online),
        "offline": len(offline),
        "abnormal": len(abnormal),
        "online_list": online,
        "offline_list": offline,
        "abnormal_list": abnormal,
        "large_offline_warning": len(offline) / total > 0.5 if total > 0 else False,
        "all_failed": total > 0 and len(online) == 0,
    }


# ─── Step 4: 更新 Excel ─────────────────────────────────

def update_excel(excel_path: str, results: list, today_str: str = None):
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl 未安装，跳过 Excel 更新")
        return False

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        target_date = today_str or date.today().strftime("%Y%m%d")

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col = headers.index(target_date) + 1 if target_date in headers else ws.max_column + 1
        if target_date not in headers:
            ws.cell(1, col, target_date)

        status_cn = {"online": "在线", "offline": "离线", "abnormal": "异常"}
        name_map = {r.get("cameraName"): r for r in results}
        code_map = {r.get("cameraCode"): r for r in results}

        for row in range(2, ws.max_row + 1):
            cam_code = ws.cell(row, 1).value
            cam_name = ws.cell(row, 2).value
            r = name_map.get(cam_name) or code_map.get(cam_code)
            if r:
                ws.cell(row, col, status_cn.get(r["status"], "未知"))

        wb.save(excel_path)
        print(f"Excel 已更新: {excel_path} (列 {target_date})")
        return True
    except Exception as e:
        print(f"Excel 更新失败: {e}")
        return False


# ─── Step 4b: 生成每日快照 Excel ─────────────────────

def generate_daily_snapshot_excel(cameras: list, results: list, config: dict,
                                   branding: dict, today_display: str):
    """
    生成『模板格式』的每日巡检快照 Excel。

    两种模式：
      1) 指定 snapshot_template（港区场景）：以参考模板为基底，保留其原有
         样式/公式/摄像头清单，仅刷新【巡检时间】列与【在线状态/离线原因】
         （按摄像头名称匹配）。所属区域、监控点名称列保持模板原样不动。
      2) 未指定模板（环保/通用场景）：从零构建 5 列表格。

    文件名:    {region_label}设备巡检记录{date}.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        print("WARNING: openpyxl 未安装，跳过快照Excel生成")
        return None

    region_label = config.get("region_label") or branding.get("region_label") or "智慧港区小脑"
    status_cn = {"online": "在线", "offline": "离线", "abnormal": "异常"}
    name_map = {r.get("cameraName"): r for r in results}
    code_map = {r.get("cameraCode"): r for r in results}

    # 港区场景：沿用参考模板样式，仅刷新时间列与离线状态
    template_path = config.get("snapshot_template")
    if template_path and os.path.exists(template_path):
        return _generate_snapshot_from_template(
            template_path, results, config, region_label, status_cn, name_map, today_display)

    # 通用场景：从零构建
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡检记录"

    # 标题行
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{today_display}巡检记录"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = ["所属区域", "监控点名称", "在线状态", "巡检时间", "离线原因"]
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(2, i, h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # 数据行
    r = 3
    for cam in cameras:
        name = cam.get("name", "")
        region = cam.get("region") or region_label
        res = code_map.get(cam.get("code")) or name_map.get(name) or {}
        status = res.get("status", "offline")
        is_online_like = status in ("online", "abnormal")
        reason = "--" if is_online_like else (res.get("errorMsg") or "未知")
        capture_time = res.get("captureTime") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        vals = [region, name, status_cn.get(status, "未知"), capture_time, reason]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(r, i, v)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            if i == 3:  # 在线状态着色
                if status == "online":
                    c.font = Font(color="006100", bold=True)
                elif status == "offline":
                    c.font = Font(color="9C0006", bold=True)
                else:
                    c.font = Font(color="9C5700", bold=True)
        r += 1

    # 列宽
    for i, w in enumerate([16, 38, 12, 22, 34], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 保存
    report_dir = config.get("report_dir") or config.get("save_dir") or "./reports"
    os.makedirs(report_dir, exist_ok=True)
    out_name = f"{region_label}设备巡检记录{today_display}.xlsx"
    out_path = os.path.join(report_dir, out_name)
    wb.save(out_path)
    print(f"快照Excel已生成: {out_path}")
    return out_path


def _generate_snapshot_from_template(template_path: str, results: list, config: dict,
                                     region_label: str, status_cn: dict,
                                     name_map: dict, today_display: str):
    """
    以参考模板为基底生成快照：保留模板全部样式与摄像头清单，
    仅刷新【巡检时间】列与【在线状态/离线原因】（按摄像头名称匹配）。
    所属区域、监控点名称列（可能是 ="..." 公式）保持原样不动。
    """
    import openpyxl
    from copy import copy as shallow_copy

    wb = openpyxl.load_workbook(template_path)   # 保留原样式 + 公式
    ws = wb.active

    # 按表头文本定位列（兼容列顺序微调）
    col = {"region": 1, "name": 2, "status": 3, "time": 4, "reason": 5}
    header_row = 2
    for hr in range(1, 4):
        vals = [ws.cell(hr, c).value for c in range(1, ws.max_column + 1)]
        if any(isinstance(v, str) and "监控点名称" in v for v in vals):
            header_row = hr
            for c in range(1, ws.max_column + 1):
                v = ws.cell(hr, c).value
                if not isinstance(v, str):
                    continue
                if "所属区域" in v: col["region"] = c
                elif "监控点名称" in v: col["name"] = c
                elif "在线状态" in v: col["status"] = c
                elif "离线原因" in v: col["reason"] = c
                elif "巡检时间" in v or "时间" in v: col["time"] = c
            break

    # 标题行日期刷新（仅改文字，保留合并/字体样式）
    title_cell = ws.cell(1, 1)
    if title_cell.value:
        title_cell.value = f"{today_display}巡检记录"

    def unwrap(v):
        if v is None:
            return ""
        s = str(v)
        if s.startswith("="):
            s = s[1:]
            if len(s) >= 2 and s[0] in "\"'":
                s = s[1:-1]
        return s.strip()

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    updated = 0
    for r in range(header_row + 1, ws.max_row + 1):
        name = unwrap(ws.cell(r, col["name"]).value)
        if not name:
            continue
        res = name_map.get(name)
        if res is None:
            continue  # 模板有但本次未巡检：保持原值不动
        status = res.get("status", "offline")
        is_online_like = status in ("online", "abnormal")
        reason = "--" if is_online_like else (res.get("errorMsg") or "未知")
        capture_time = res.get("captureTime") or now_str
        # 仅覆盖 时间列 / 状态列 / 离线原因列
        ws.cell(r, col["time"]).value = capture_time
        ws.cell(r, col["status"]).value = status_cn.get(status, "未知")
        ws.cell(r, col["reason"]).value = reason
        updated += 1

    report_dir = config.get("report_dir") or config.get("save_dir") or "./reports"
    os.makedirs(report_dir, exist_ok=True)
    out_name = f"{region_label}设备巡检记录{today_display}.xlsx"
    out_path = os.path.join(report_dir, out_name)
    wb.save(out_path)
    print(f"快照Excel(沿用模板样式)已生成: {out_path}  (刷新 {updated} 路)")
    return out_path


# ─── Step 5: 飞书通知 ─────────────────────────────────

def send_feishu_notification(summary: dict, webhook_url: str, branding: dict = None):
    if not webhook_url:
        print("[飞书] 未配置 webhook_url，跳过通知")
        return False

    import requests

    if branding is None:
        branding = {}

    card_title = branding.get("feishu_card_title", "摄像头巡检报告")
    card_footer = branding.get("feishu_card_footer", "巡检助手 · 自动生成")
    inspection_banner = branding.get("inspection_start_banner", "摄像头巡检")

    offline_text = "\n".join(
        f"❌ {r['cameraName']}：{r.get('errorMsg', '超时')}" for r in summary["offline_list"]
    ) or "无"
    abnormal_text = "\n".join(
        f"⚠️ {r['cameraName']}：{r.get('errorMsg', '画面质量差')}" for r in summary["abnormal_list"]
    ) or "无"

    alert = ""
    if summary["all_failed"]:
        alert = "\n🚨 **全部摄像头离线**：可能是海康平台停机或网络中断！"
    elif summary["large_offline_warning"]:
        alert = "\n⚠️ **大面积离线警告**：超过50%摄像头离线，可能是平台或网络问题！"

    body = {
        "msg_type": "interactive",
        "card": {
            "header": {"title": {"tag": "plain_text", "content": card_title}, "template": "blue"},
            "elements": [
                {"tag": "div", "text": {"tag": "lark_md", "content": (
                    f"**执行时间**: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                    f"**总计**: {summary['total']} 路  "
                    f"**✅ 在线**: {summary['online']}  "
                    f"**❌ 离线**: {summary['offline']}  "
                    f"**⚠️ 异常**: {summary['abnormal']}"
                    f"{alert}"
                )}},
                {"tag": "hr"},
                {"tag": "div", "text": {"tag": "lark_md", "content": f"**❌ 离线摄像头 ({summary['offline']})：**\n{offline_text}"}},
                {"tag": "hr"},
                {"tag": "div", "text": {"tag": "lark_md", "content": f"**⚠️ 异常摄像头 ({summary['abnormal']})：**\n{abnormal_text}"}},
                {"tag": "hr"},
                {"tag": "note", "elements": [{"tag": "plain_text", "content": card_footer}]},
            ],
        },
    }

    try:
        resp = requests.post(webhook_url, json=body, timeout=10)
        resp.raise_for_status()
        result = resp.json()
        print(f"[飞书] 发送结果: {result}")
        return result.get("StatusCode") == 0
    except Exception as e:
        print(f"[飞书] 发送失败: {e}")
        return False


# ─── Step 5: 更新 Word 巡查台账 ──────────────────────

def update_docx_ledger(results: list, config: dict, today_str: str) -> bool:
    """
    将巡检截图贴入 Word 巡查台账。

    仅处理 Excel 中"是否更新到台账"列为"是"的摄像头，
    运营人员只需在 Excel 中维护该列即可控制哪些摄像头写入台账。

    config 中的相关字段：
      docx_template  — Word 模版路径（必填，否则跳过本步骤）
      docx_output    — 输出 Word 路径（可选，默认在模版同目录下按日期命名）
      docx_img_width — 图片宽度（英寸），默认 1.8
    """
    docx_template = config.get("docx_template", "")
    if not docx_template:
        print("[台账] 未配置 docx_template，跳过 Word 台账更新")
        return False

    if not os.path.exists(docx_template):
        print(f"[台账] 模版文件不存在: {docx_template}，跳过")
        return False

    # 按 Excel "是否更新到台账" 列过滤：仅处理 ledgerEnabled=True 的摄像头
    ledger_results = [r for r in results if r.get("ledgerEnabled", False)]
    skipped = len(results) - len(ledger_results)
    if skipped > 0:
        print(f"[台账] 按 Excel 配置过滤：{len(ledger_results)} 路写入台账，{skipped} 路跳过")
    if not ledger_results:
        print("[台账] 无需要写入台账的摄像头，跳过")
        return False

    # 输出路径默认：与模版同目录，按日期命名
    docx_output = config.get("docx_output", "")
    if not docx_output:
        template_dir = os.path.dirname(os.path.abspath(docx_template))
        template_base = os.path.splitext(os.path.basename(docx_template))[0]
        docx_output = os.path.join(template_dir, f"{template_base}_{today_str}.docx")

    img_width = float(config.get("docx_img_width", 1.8))

    # 动态导入（脚本与本文件同目录）
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, script_dir)
    try:
        from update_docx import update_docx
    except ImportError as e:
        print(f"[台账] 无法导入 update_docx 模块: {e}，跳过")
        return False

    # 日期格式：YYYY.M.D（去掉补零，符合文档习惯）
    d = date.today()
    today_display = f"{d.year}.{d.month}.{d.day}"

    print(f"  模版: {docx_template}")
    print(f"  输出: {docx_output}")

    try:
        stat = update_docx(
            docx_path=docx_template,
            results=ledger_results,
            output_path=docx_output,
            today_str=today_display,
            img_width_inches=img_width,
            skip_offline=True,
        )
        return stat["matched"] > 0
    except Exception as e:
        print(f"[台账] 更新失败: {e}")
        return False


# ─── 主流程 ──────────────────────────────────────────

def run_inspection(config: dict):
    """
    执行完整巡检流程。
    参数 config 已由外层合并好：平台凭证 + 运行时配置。
    """
    host = config.get("host") or os.environ.get("HIK_HOST", "")
    app_key = config.get("app_key") or os.environ.get("HIK_APP_KEY", "")
    app_secret = config.get("app_secret") or os.environ.get("HIK_APP_SECRET", "")

    if not host or not app_key or not app_secret:
        print("ERROR: 缺少必填配置（host/app_key/app_secret）")
        print("请检查技能包 config/platform.json 或设置环境变量 HIK_HOST/HIK_APP_KEY/HIK_APP_SECRET")
        sys.exit(1)

    config["host"] = host
    config["app_key"] = app_key
    config["app_secret"] = app_secret
    config.setdefault("port", 443)
    config.setdefault("save_dir", "./screenshots")
    config.setdefault("timeout", 10)
    config.setdefault("retry_count", 2)
    config.setdefault("max_workers", 30)
    config.setdefault("api_path", "/artemis")

    excel_path = config.get("excel_path", "")
    if not excel_path or not os.path.exists(excel_path):
        print(f"ERROR: 摄像头清单文件未提供或不存在: {excel_path}")
        print("请通过 --excel-path 参数指定摄像头清单路径")
        sys.exit(1)

    # 加载品牌配置（港区场景会从合并后的 config["branding"] 读取）
    branding = load_branding_config(config)
    banner = branding.get("inspection_start_banner", "摄像头巡检")

    print(f"\n{'='*50}")
    print(f"  {banner}开始")
    print(f"  平台: {config['host']}:{config['port']}")
    print(f"  清单: {excel_path}")
    print(f"  截图: {config['save_dir']}")
    print(f"{'='*50}\n")

    # Step 1: 读取清单
    print("[Step 1] 读取摄像头清单...")
    cameras = parse_camera_excel(excel_path)
    print(f"  共 {len(cameras)} 个摄像头")
    if not cameras:
        print("ERROR: 未解析到任何摄像头")
        sys.exit(1)

    # Step 2: 并发截图
    print(f"\n[Step 2] 并发截图巡检...")
    inspection_start = time.time()
    results = execute_concurrent_inspection(cameras, config)
    elapsed = time.time() - inspection_start
    print(f"  耗时 {elapsed:.1f}s")

    # Step 3: 汇总
    print(f"\n[Step 3] 汇总结果...")
    summary = summarize(results)
    print(f"  总计: {summary['total']}  |  在线: {summary['online']}  |  离线: {summary['offline']}  |  异常: {summary['abnormal']}")
    if summary["large_offline_warning"]:
        print("  ⚠️  大面积离线警告!")
    if summary["all_failed"]:
        print("  🚨 全部摄像头离线!")

    # Step 4: 更新 Excel（按模式分流）
    today_str = date.today().strftime("%Y%m%d")
    today_display = date.today().strftime("%Y-%m-%d")
    excel_mode = config.get("excel_mode", "accumulate")
    if excel_mode == "snapshot":
        print(f"\n[Step 4] 生成每日快照Excel...")
        generate_daily_snapshot_excel(cameras, results, config, branding, today_display)
    else:
        print(f"\n[Step 4] 更新 Excel（累计模式）...")
        update_excel(excel_path, results, today_str)

    # Step 5: 更新 Word 巡查台账
    print(f"\n[Step 5] 更新 Word 巡查台账...")
    update_docx_ledger(results, config, today_str)

    # Step 6: 飞书通知
    print(f"\n[Step 6] 发送飞书通知...")
    feishu_webhook = config.get("feishu_webhook_url") or os.environ.get("FEISHU_WEBHOOK_URL", "")
    send_feishu_notification(summary, feishu_webhook, branding)

    # 保存 JSON 结果
    output_dir = config["save_dir"]
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"inspection_{today_str}.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n结果已保存: {output_path}")

    print(f"\n{'='*50}")
    print(f"  巡检完成! 总耗时: {elapsed:.1f}s")
    print(f"{'='*50}")

    return 1 if summary["offline"] > 0 or summary["abnormal"] > 0 else 0


def main():
    import argparse
    parser = argparse.ArgumentParser(description="摄像头巡检（品牌信息从 config/branding.json 读取）")
    parser.add_argument("--config", default=None, help="完全自定义配置文件路径（覆盖技能包本地配置）")
    parser.add_argument("--host", default=None, help="海康平台地址")
    parser.add_argument("--port", type=int, default=None, help="HTTPS 端口（默认 443）")
    parser.add_argument("--app-key", default=None, help="海康平台 appKey")
    parser.add_argument("--app-secret", default=None, help="海康平台 appSecret")
    parser.add_argument("--excel-path", default=None, help="摄像头清单 Excel 文件路径")
    parser.add_argument("--save-dir", default=None, help="截图保存目录")
    parser.add_argument("--feishu-webhook", default=None, help="飞书机器人 Webhook URL（覆盖配置中的值）")
    parser.add_argument("--docx-template", default=None, help="Word 巡查台账模版路径（如: ./巡查台账_模版.docx）")
    parser.add_argument("--docx-output", default=None, help="Word 台账输出路径（默认按日期自动命名）")
    parser.add_argument("--excel-mode", default=None, choices=["accumulate", "snapshot"],
                        help="Excel 输出模式：accumulate=累计追加列(环保默认) / snapshot=每日快照模板(港区)")
    parser.add_argument("--report-dir", default=None, help="snapshot 模式下的快照Excel输出目录")
    parser.add_argument("--region-label", default=None, help="snapshot 模式下『所属区域』列 / 文件标题前缀")
    parser.add_argument("--snapshot-template", default=None, help="snapshot 模式参考模板路径（沿用其样式，仅刷新时间列与离线状态）")
    parser.add_argument("--test-single", default=None, help="测试单个摄像头 (传编码)")
    args = parser.parse_args()

    if args.test_single:
        # 单路测试模式：从技能包本地配置加载平台凭证
        config = load_platform_config()
        if args.config and os.path.exists(args.config):
            with open(args.config, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                config.update(cfg)

        host = config.get("host") or os.environ.get("HIK_HOST", "")
        app_key = config.get("app_key") or os.environ.get("HIK_APP_KEY", "")
        app_secret = config.get("app_secret") or os.environ.get("HIK_APP_SECRET", "")
        if not host or not app_key or not app_secret:
            print("ERROR: 缺少配置")
            sys.exit(1)

        result = capture_single(
            host=host, port=int(config.get("port", 443)),
            app_key=app_key, app_secret=app_secret,
            camera_index_code=args.test_single, camera_name="",
            save_dir=config.get("save_dir", "./screenshots"),
            timeout=int(config.get("timeout", 10)),
            retry_count=int(config.get("retry_count", 2)),
            api_path=config.get("api_path", "/artemis"),
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(0 if result.get("status") == "online" else 1)

    # 加载配置：平台凭证(含密钥)始终以技能包 config/platform.json 为基准，
    # --config 仅作覆盖/补充（港区配置无需重复写密钥）。
    if args.config:
        if not os.path.exists(args.config):
            print(f"ERROR: 配置文件不存在: {args.config}")
            sys.exit(1)
        base = load_platform_config()
        with open(args.config, "r", encoding="utf-8") as f:
            override = json.load(f)
        base.update(override)
        config = base
        print(f"[配置] 基准 platform.json + 覆盖 {args.config}")
    else:
        config = load_platform_config()

    # 合并用户运行时配置（CLI 参数优先级最高，可覆盖一切）
    if args.host:
        config["host"] = args.host
    if args.port is not None:
        config["port"] = args.port
    if args.app_key:
        config["app_key"] = args.app_key
    if args.app_secret:
        config["app_secret"] = args.app_secret
    if args.excel_path:
        config["excel_path"] = args.excel_path
    if args.save_dir:
        config["save_dir"] = args.save_dir
    if args.feishu_webhook:
        config["feishu_webhook_url"] = args.feishu_webhook
    if args.docx_template:
        config["docx_template"] = args.docx_template
    if args.docx_output:
        config["docx_output"] = args.docx_output
    if args.excel_mode:
        config["excel_mode"] = args.excel_mode
    if args.report_dir:
        config["report_dir"] = args.report_dir
    if args.region_label:
        config["region_label"] = args.region_label
    if args.snapshot_template:
        config["snapshot_template"] = args.snapshot_template

    sys.exit(run_inspection(config))


if __name__ == "__main__":
    main()
